import os
import json
import re
import base64
import requests
from datetime import datetime
from google import genai
from google.genai import types
from pypdf import PdfReader
from dotenv import load_dotenv

load_dotenv()

GEMINI_API_KEY = os.environ.get("GEMINI_API_KEY", "")
OPENAI_API_KEY = os.environ.get("OPENAI_API_KEY", "")
ANTHROPIC_API_KEY = os.environ.get("ANTHROPIC_API_KEY", "")

STANDARD_CATEGORIES = [
    "Food & Dining",
    "Housing & Rent",
    "Utilities & Bills",
    "Transportation",
    "Shopping & Retail",
    "Entertainment & Leisure",
    "Healthcare & Wellness",
    "Education & Learning",
    "Personal Care",
    "Investments & Savings",
    "Refunds & Credits",
    "Miscellaneous"
]

EXTRACTION_SYSTEM_PROMPT = f"""
You are an expert financial document analyzer.
Analyze the provided statement (Bank Statement, Amex credit card bill, Excel spreadsheet, invoice, or receipt) and extract all expense and refund transactions.

Rules for Extraction:
1. Extract every purchase, debit, merchant fee, and merchant return/statement credit.
2. HANDLING CREDITS VS BILL PAYMENTS:
   - MERCHANT REFUNDS, RETURNS & CREDITS (e.g. Amazon refund, Airline cancellation credit, merchant dispute credit, cash back rewards): DO EXTRACT THESE. Classify them under 'Refunds & Credits' or the original purchase category, and clearly note '(Refund / Credit)' in description (e.g. "Amazon.com (Refund Credit)").
   - CREDIT CARD BILL PAYMENTS & SETTLEMENTS (e.g. "AUTOPAY PAYMENT RECEIVED - THANK YOU", "ONLINE PAYMENT RECEIVED", "DIRECT DEBIT PYMT"): DO NOT INCLUDE THESE (these are balance pay-downs from bank checking accounts, not purchases or merchant returns).
3. Standardize the Date into 'YYYY-MM-DD' format.
4. Extract the exact numerical Amount as a positive float (e.g. 45.20).
5. Clean up the Merchant / Description name (e.g. change "WHOLEFDS MKT 1024 SAN FRANCISCO CA" to "Whole Foods Market").
6. Classify each transaction into EXACTLY ONE of the following categories:
   {json.dumps(STANDARD_CATEGORIES)}
7. Identify Payment Method (e.g., "Amex Card", "Credit Card", "Debit Card", "Bank Transfer", "Cash", "UPI / Online").
8. Return ONLY a valid JSON array of objects with the exact schema:
[
  {{
    "date": "YYYY-MM-DD",
    "amount": 25.50,
    "category": "Food & Dining",
    "payment_method": "Amex Card",
    "description": "Starbucks Coffee"
  }}
]
"""

def normalize_date_string(date_raw):
    if not date_raw:
        return datetime.now().strftime("%Y-%m-%d")
    date_str = str(date_raw).strip().split('T')[0].split(' ')[0]
    for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%d/%m/%Y", "%Y/%m/%d", "%m-%d-%Y", "%d-%m-%Y", "%b %d, %Y", "%B %d, %Y", "%d %b %Y", "%d-%b-%Y"):
        try:
            dt = datetime.strptime(date_str, fmt)
            return dt.strftime("%Y-%m-%d")
        except Exception:
            pass
    match = re.search(r'(\d{4})[-/](\d{1,2})[-/](\d{1,2})', str(date_raw))
    if match:
        y, m, d = match.groups()
        return f"{int(y):04d}-{int(m):02d}-{int(d):02d}"
    match2 = re.search(r'(\d{1,2})[-/](\d{1,2})[-/](\d{4})', str(date_raw))
    if match2:
        m, d, y = match2.groups()
        return f"{int(y):04d}-{int(m):02d}-{int(d):02d}"
    return datetime.now().strftime("%Y-%m-%d")

def extract_text_from_pdf(file_stream_or_path):
    try:
        if hasattr(file_stream_or_path, 'seek'):
            file_stream_or_path.seek(0)
        reader = PdfReader(file_stream_or_path)
        text = ""
        for i, page in enumerate(reader.pages):
            page_text = page.extract_text()
            if page_text:
                text += f"\n--- Page {i+1} ---\n" + page_text
        return text
    except Exception as e:
        print(f"Error reading PDF: {e}")
        return ""

def extract_text_from_excel(file_stream_or_path):
    try:
        if hasattr(file_stream_or_path, 'seek'):
            file_stream_or_path.seek(0)
        import openpyxl
        wb = openpyxl.load_workbook(file_stream_or_path, data_only=True)
        text_lines = []
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            text_lines.append(f"--- Sheet: {sheet_name} ---")
            for row in ws.iter_rows(values_only=True):
                if any(v is not None for v in row):
                    line = " | ".join(str(v).strip() for v in row if v is not None)
                    text_lines.append(line)
        return "\n".join(text_lines)
    except Exception as e:
        print(f"Error reading Excel via openpyxl: {e}")
        try:
            if hasattr(file_stream_or_path, 'seek'):
                file_stream_or_path.seek(0)
            import pandas as pd
            xls = pd.ExcelFile(file_stream_or_path)
            text_lines = []
            for sheet in xls.sheet_names:
                df = pd.read_excel(xls, sheet_name=sheet)
                text_lines.append(f"--- Sheet: {sheet} ---")
                text_lines.append(df.to_string(index=False))
            return "\n".join(text_lines)
        except Exception as e2:
            print(f"Error reading Excel via pandas: {e2}")
            return ""

def clean_json_response(raw_text):
    if not raw_text:
        return []
    cleaned = raw_text.strip()
    if cleaned.startswith("```"):
        cleaned = re.sub(r"^```(?:json)?\n?", "", cleaned)
        cleaned = re.sub(r"\n?```$", "", cleaned)
        cleaned = cleaned.strip()
    try:
        return json.loads(cleaned)
    except Exception:
        pass
    array_match = re.search(r'\[.*\]', cleaned, re.DOTALL)
    if array_match:
        try:
            return json.loads(array_match.group(0))
        except Exception:
            pass
    obj_match = re.search(r'\{.*\}', cleaned, re.DOTALL)
    if obj_match:
        try:
            return json.loads(obj_match.group(0))
        except Exception:
            pass
    return []

def get_active_ai_credentials(custom_api_key=None, custom_provider=None):
    key = (custom_api_key or "").strip()
    provider = (custom_provider or "").strip().lower()

    if not key:
        key = os.environ.get("GEMINI_API_KEY", "") or os.environ.get("OPENAI_API_KEY", "") or os.environ.get("ANTHROPIC_API_KEY", "")
        if not key:
            cfg_file = os.path.join(os.path.dirname(os.path.abspath(__file__)), ".api_key")
            if os.path.exists(cfg_file):
                try:
                    with open(cfg_file, "r") as f:
                        key = f.read().strip()
                except Exception:
                    pass

    if not provider:
        if key.startswith("sk-ant-"):
            provider = "anthropic"
        elif key.startswith("sk-or-"):
            provider = "openrouter"
        elif key.startswith("sk-"):
            provider = "openai"
        else:
            provider = "gemini"

    return key, provider

def execute_ai_completion(prompt, system_instruction=None, image_bytes=None, mime_type=None, custom_api_key=None, custom_provider=None):
    api_key, provider = get_active_ai_credentials(custom_api_key, custom_provider)
    if not api_key:
        raise ValueError("AI API key is not configured. Please set your key in Settings or environment variables.")

    if provider == "openai":
        headers = {
            "Authorization": f"Bearer {api_key}",
            "Content-Type": "application/json"
        }
        messages = []
        if system_instruction:
            messages.append({"role": "system", "content": system_instruction})

        if image_bytes:
            b64_img = base64.b64encode(image_bytes).decode('utf-8')
            img_url = f"data:{mime_type or 'image/png'};base64,{b64_img}"
            messages.append({
                "role": "user",
                "content": [
                    {"type": "text", "text": prompt},
                    {"type": "image_url", "image_url": {"url": img_url}}
                ]
            })
        else:
            messages.append({"role": "user", "content": prompt})

        resp = requests.post(
            "https://api.openai.com/v1/chat/completions",
            headers=headers,
            json={"model": "gpt-4o", "messages": messages, "temperature": 0.2},
            timeout=60
        )
        if resp.status_code != 200:
            raise ValueError(f"OpenAI API Error ({resp.status_code}): {resp.text}")
        data = resp.json()
        return data["choices"][0]["message"]["content"]

    elif provider in ["anthropic", "claude"]:
        headers = {
            "x-api-key": api_key,
            "anthropic-version": "2023-06-01",
            "content-type": "application/json"
        }
        user_content = []
        if image_bytes:
            b64_img = base64.b64encode(image_bytes).decode('utf-8')
            user_content.append({
                "type": "image",
                "source": {
                    "type": "base64",
                    "media_type": mime_type or "image/png",
                    "data": b64_img
                }
            })
        user_content.append({"type": "text", "text": prompt})

        payload = {
            "model": "claude-3-5-sonnet-20241022",
            "max_tokens": 4096,
            "messages": [{"role": "user", "content": user_content}]
        }
        if system_instruction:
            payload["system"] = system_instruction

        resp = requests.post(
            "https://api.anthropic.com/v1/messages",
            headers=headers,
            json=payload,
            timeout=60
        )
        if resp.status_code != 200:
            raise ValueError(f"Anthropic API Error ({resp.status_code}): {resp.text}")
        data = resp.json()
        return data["content"][0]["text"]

    elif provider == "openrouter":
        headers = {
            "Authorization": f"Bearer {api_key}",
            "Content-Type": "application/json"
        }
        messages = []
        if system_instruction:
            messages.append({"role": "system", "content": system_instruction})
        messages.append({"role": "user", "content": prompt})

        resp = requests.post(
            "https://openrouter.ai/api/v1/chat/completions",
            headers=headers,
            json={"model": "anthropic/claude-3.5-sonnet", "messages": messages},
            timeout=60
        )
        if resp.status_code != 200:
            raise ValueError(f"OpenRouter API Error ({resp.status_code}): {resp.text}")
        data = resp.json()
        return data["choices"][0]["message"]["content"]

    else:
        client = genai.Client(api_key=api_key)
        models_to_try = ["gemini-flash-latest", "gemini-flash-lite-latest", "gemini-3.5-flash", "gemini-3.1-flash-lite"]
        full_prompt = f"{system_instruction}\n\n{prompt}" if system_instruction else prompt
        last_error = None

        if image_bytes:
            image_part = types.Part.from_bytes(data=image_bytes, mime_type=mime_type or "image/png")
            for m in models_to_try:
                try:
                    chat = client.chats.create(model=m)
                    res = chat.send_message([full_prompt, image_part])
                    if res.text: return res.text
                except Exception as err:
                    last_error = err
                    continue
        else:
            for m in models_to_try:
                try:
                    chat = client.chats.create(model=m)
                    res = chat.send_message(full_prompt)
                    if res.text: return res.text
                except Exception as err:
                    last_error = err
                    continue

        raise ValueError(f"Gemini API Error: {last_error}")

def parse_statement(file_obj, filename="", custom_api_key=None, custom_provider=None):
    ext = os.path.splitext(filename.lower())[1] if filename else ""
    raw_response_text = ""

    if ext == ".pdf":
        extracted_text = extract_text_from_pdf(file_obj)
        if not extracted_text or len(extracted_text.strip()) < 10:
            raise ValueError("Could not extract readable text from PDF statement.")
        prompt = f"{EXTRACTION_SYSTEM_PROMPT}\n\nHere is the raw text from the statement:\n\n{extracted_text}"
        raw_response_text = execute_ai_completion(prompt, custom_api_key=custom_api_key, custom_provider=custom_provider)

    elif ext in [".xlsx", ".xls"]:
        extracted_text = extract_text_from_excel(file_obj)
        if not extracted_text or len(extracted_text.strip()) < 10:
            raise ValueError("Could not extract readable tabular data from the Excel spreadsheet.")
        prompt = f"{EXTRACTION_SYSTEM_PROMPT}\n\nHere is the data from the Excel file ({filename}):\n\n{extracted_text}"
        raw_response_text = execute_ai_completion(prompt, custom_api_key=custom_api_key, custom_provider=custom_provider)

    elif ext in [".csv", ".txt"]:
        if hasattr(file_obj, 'seek'):
            file_obj.seek(0)
        content = file_obj.read()
        if isinstance(content, bytes):
            content = content.decode('utf-8', errors='ignore')
        prompt = f"{EXTRACTION_SYSTEM_PROMPT}\n\nHere is the statement content ({filename}):\n\n{content}"
        raw_response_text = execute_ai_completion(prompt, custom_api_key=custom_api_key, custom_provider=custom_provider)

    elif ext in [".png", ".jpg", ".jpeg", ".webp"]:
        if hasattr(file_obj, 'seek'):
            file_obj.seek(0)
        file_bytes = file_obj.read()
        mime_type = "image/png" if ext == ".png" else "image/jpeg" if ext in [".jpg", ".jpeg"] else "image/webp"
        raw_response_text = execute_ai_completion(
            EXTRACTION_SYSTEM_PROMPT,
            image_bytes=file_bytes,
            mime_type=mime_type,
            custom_api_key=custom_api_key,
            custom_provider=custom_provider
        )
    else:
        if hasattr(file_obj, 'seek'):
            file_obj.seek(0)
        content = file_obj.read()
        if isinstance(content, bytes):
            content = content.decode('utf-8', errors='ignore')
        prompt = f"{EXTRACTION_SYSTEM_PROMPT}\n\nStatement Content:\n\n{content}"
        raw_response_text = execute_ai_completion(prompt, custom_api_key=custom_api_key, custom_provider=custom_provider)

    parsed_transactions = clean_json_response(raw_response_text)
    
    valid_items = []
    if isinstance(parsed_transactions, list):
        for item in parsed_transactions:
            if not isinstance(item, dict):
                continue
            try:
                amt = float(item.get("amount", 0))
                if amt <= 0:
                    continue
                cat = str(item.get("category", "Miscellaneous")).strip()
                if cat not in STANDARD_CATEGORIES:
                    cat = "Miscellaneous"
                
                date_str = normalize_date_string(item.get("date"))
                desc = str(item.get("description", "Purchase")).strip()
                pay = str(item.get("payment_method", "Amex Card")).strip()
                
                valid_items.append({
                    "date": date_str,
                    "amount": round(amt, 2),
                    "category": cat,
                    "payment_method": pay,
                    "description": desc
                })
            except Exception:
                continue

    return valid_items

FINANCIAL_INSIGHTS_PROMPT = """
You are an expert personal financial advisor and wealth coach.
Analyze the user's spending data, monthly budget, category allocations, and recent expenses.

Generate clear, engaging, and actionable financial insights.
Requirements:
1. "health_score": An integer from 0 to 100 representing their budget health (100 is best).
2. "status": A concise status word (e.g. "Excellent", "Healthy", "Caution", "Critical Overspending").
3. "headline": A punchy 1-sentence executive summary of their current financial performance.
4. "observations": Array of 2 to 3 data-driven observations highlighting top spending categories, daily burn rate, or notable patterns.
5. "recommendations": Array of 3 specific, actionable recommendations to reduce expenses and save money.
6. "alerts": Array of warning strings for categories that exceeded or are nearing their budget limit (empty array if none).
7. "projected_monthly_savings": A realistic estimated dollar amount ($) they could save by following your tips.

Return ONLY a valid JSON object matching this schema:
{
  "health_score": 85,
  "status": "Healthy",
  "headline": "Great control over fixed costs, with minor savings potential in Dining.",
  "observations": [
    "Your top expense category this month is Food & Dining ($320).",
    "You have used 48% of your overall monthly budget limit."
  ],
  "recommendations": [
    "Prepare meals at home 2 more days per week to save ~$120/mo.",
    "Review recurring subscription renewals in Entertainment."
  ],
  "alerts": [
    "Dining spend has reached 85% of its $400 budget allowance."
  ],
  "projected_monthly_savings": 140.00
}
"""

def generate_financial_insights(month_label, summary_data, expenses_data, custom_api_key=None, custom_provider=None):
    payload = {
        "period": month_label,
        "total_spent": summary_data.get("current_month_spend", 0.0),
        "total_budget": summary_data.get("total_monthly_budget", 0.0),
        "remaining_budget": summary_data.get("remaining_budget", 0.0),
        "budget_usage_pct": summary_data.get("budget_usage_pct", 0.0),
        "daily_average": summary_data.get("daily_avg_spend", 0.0),
        "category_breakdown": summary_data.get("category_breakdown", []),
        "budget_comparison": summary_data.get("budget_comparison", []),
        "recent_expenses": expenses_data[:20] if expenses_data else []
    }

    prompt = f"{FINANCIAL_INSIGHTS_PROMPT}\n\nHere is the financial data for {month_label}:\n\n{json.dumps(payload, indent=2)}"
    raw_response_text = execute_ai_completion(prompt, custom_api_key=custom_api_key, custom_provider=custom_provider)

    text = str(raw_response_text or '').strip()
    if text.startswith("```"):
        text = re.sub(r"^```(?:json)?\n?", "", text)
        text = re.sub(r"\n?```$", "", text)

    try:
        data = json.loads(text)
        if isinstance(data, dict) and "health_score" in data:
            return data
    except Exception:
        pass

    return {
        "health_score": 75,
        "status": "Healthy",
        "headline": f"Financial overview generated for {month_label}.",
        "observations": [
            f"Total spend for {month_label} is ${summary_data.get('current_month_spend', 0.0):.2f}.",
            f"Budget utilization is at {summary_data.get('budget_usage_pct', 0.0)}%."
        ],
        "recommendations": [
            "Continue tracking daily expenses to maintain budget limits.",
            "Review categories nearing 80% utilization."
        ],
        "alerts": [],
        "projected_monthly_savings": 50.00
    }

COPILOT_SYSTEM_PROMPT = """
You are "Expenz Copilot", an elite, proactive, and friendly personal AI financial advisor built directly into Expenz.io.
You have direct, real-time access to the user's active Excel ledger, category budgets, expense transactions, and spending statistics.

Your Persona & Rules:
1. Provide accurate, insightful, and grounded answers using the user's actual expense data provided below.
2. Quote exact numbers ($ amounts, merchant names, dates, budget utilization %) whenever applicable.
3. Format your response in clean, beautiful GitHub-flavored markdown with bold highlights, clean bullet points, and easy-to-read sections.
4. When giving financial advice, be encouraging, pragmatic, and actionable (e.g. identify specific subscriptions or high dining weeks to cut).
5. Provide 2 to 3 concise, highly relevant follow-up questions the user might want to click next.
6. Always return ONLY a valid JSON object with the exact structure:
{
  "reply": "Your markdown formatted reply here...",
  "suggested_followups": [
    "What was my largest purchase this month?",
    "How can I save $150 on dining?"
  ]
}
"""

def generate_copilot_response(user_message, history, summary_data, all_expenses, custom_api_key=None, custom_provider=None):
    period = summary_data.get("active_month_label", "Current Overview")
    total_spent = summary_data.get("current_month_spend", 0.0)
    total_budget = summary_data.get("total_monthly_budget", 0.0)
    remaining = summary_data.get("remaining_budget", 0.0)
    budget_usage = summary_data.get("budget_usage_pct", 0.0)
    cat_breakdown = summary_data.get("category_breakdown", [])
    budget_comp = summary_data.get("budget_comparison", [])
    recent_txns = all_expenses[:60] if all_expenses else []

    context_payload = {
        "active_period": period,
        "total_spent": total_spent,
        "total_budget": total_budget,
        "remaining_budget": remaining,
        "budget_usage_percent": budget_usage,
        "category_spending": cat_breakdown,
        "category_budgets": budget_comp,
        "recent_transactions_sample": recent_txns,
        "total_transactions_count": len(all_expenses)
    }

    conversation_text = ""
    if history and isinstance(history, list):
        for msg in history[-8:]:
            role = "User" if msg.get("role") == "user" else "Copilot"
            conversation_text += f"{role}: {msg.get('text', '')}\n"

    prompt = f"""=== USER'S LIVE FINANCIAL LEDGER DATA ===
{json.dumps(context_payload, indent=2)}

=== CONVERSATION HISTORY ===
{conversation_text}

User Question: {user_message}

Return ONLY the JSON response with 'reply' and 'suggested_followups':"""

    raw_response_text = execute_ai_completion(
        prompt=prompt,
        system_instruction=COPILOT_SYSTEM_PROMPT,
        custom_api_key=custom_api_key,
        custom_provider=custom_provider
    )

    text = raw_response_text.strip()
    if text.startswith("```"):
        text = re.sub(r"^```(?:json)?\n?", "", text)
        text = re.sub(r"\n?```$", "", text)

    match = re.search(r'\{.*\}', text, re.DOTALL)
    if match:
        try:
            return json.loads(match.group(0))
        except Exception:
            pass

    try:
        return json.loads(text)
    except Exception:
        return {
            "reply": text,
            "suggested_followups": [
                "What is my highest expense category?",
                "How much budget do I have left?"
            ]
        }

