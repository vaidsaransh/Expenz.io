import os
import threading
from datetime import datetime, timedelta
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
USER_DATA_DIR = os.path.join(BASE_DIR, "user_data")
EXCEL_FILE = os.path.join(BASE_DIR, "expenses_data.xlsx")
_file_lock = threading.RLock()

def get_excel_file_path(user_id=None):
    if not user_id or str(user_id).strip().lower() in ("default", "main", "primary", "", "none"):
        return EXCEL_FILE
    
    clean_id = "".join(c for c in str(user_id) if c.isalnum() or c in ("-", "_")).strip()
    if not clean_id or clean_id.lower() in ("default", "main"):
        return EXCEL_FILE
    
    clean_id = clean_id[:64]
    os.makedirs(USER_DATA_DIR, exist_ok=True)
    return os.path.join(USER_DATA_DIR, f"expenses_{clean_id}.xlsx")

DEFAULT_CATEGORIES = [
    {"name": "Food & Dining", "color": "#F59E0B", "icon": "utensils"},
    {"name": "Housing & Rent", "color": "#6366F1", "icon": "home"},
    {"name": "Utilities & Bills", "color": "#06B6D4", "icon": "bolt"},
    {"name": "Transportation", "color": "#3B82F6", "icon": "car"},
    {"name": "Shopping & Retail", "color": "#EC4899", "icon": "shopping-bag"},
    {"name": "Entertainment & Leisure", "color": "#8B5CF6", "icon": "film"},
    {"name": "Healthcare & Wellness", "color": "#10B981", "icon": "heart-pulse"},
    {"name": "Education & Learning", "color": "#14B8A6", "icon": "graduation-cap"},
    {"name": "Personal Care", "color": "#F43F5E", "icon": "sparkles"},
    {"name": "Investments & Savings", "color": "#84CC16", "icon": "piggy-bank"},
    {"name": "Refunds & Credits", "color": "#10B981", "icon": "rotate-left"},
    {"name": "Miscellaneous", "color": "#64748B", "icon": "tags"},
]

DEFAULT_BUDGETS = {
    "Food & Dining": 600.0,
    "Housing & Rent": 1500.0,
    "Utilities & Bills": 250.0,
    "Transportation": 300.0,
    "Shopping & Retail": 400.0,
    "Entertainment & Leisure": 250.0,
    "Healthcare & Wellness": 200.0,
    "Education & Learning": 150.0,
    "Personal Care": 150.0,
    "Investments & Savings": 500.0,
    "Miscellaneous": 100.0,
}

def style_header(sheet, headers, fill_color="1E293B", text_color="FFFFFF"):
    font = Font(name="Calibri", size=11, bold=True, color=text_color)
    fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
    align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style='thin', color='CBD5E1'),
        right=Side(style='thin', color='CBD5E1'),
        top=Side(style='thin', color='CBD5E1'),
        bottom=Side(style='thin', color='CBD5E1')
    )
    sheet.append(headers)
    for col_num, _ in enumerate(headers, 1):
        cell = sheet.cell(row=1, column=col_num)
        cell.font = font
        cell.fill = fill
        cell.alignment = align
        cell.border = thin_border
    sheet.row_dimensions[1].height = 26

def auto_fit_columns(sheet):
    for col in sheet.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            val_str = str(cell.value or '')
            if len(val_str) > max_len:
                max_len = len(val_str)
        sheet.column_dimensions[col_letter].width = max(max_len + 4, 12)

def init_excel(target_file=None):
    if target_file is None:
        target_file = EXCEL_FILE
    with _file_lock:
        if os.path.exists(target_file):
            return target_file

        # If creating a user workspace and master template exists, seed from master
        if target_file != EXCEL_FILE and os.path.exists(EXCEL_FILE):
            try:
                import shutil
                shutil.copyfile(EXCEL_FILE, target_file)
                return target_file
            except Exception:
                pass

        wb = openpyxl.Workbook()
        
        # 1. Expenses Sheet
        ws_expenses = wb.active
        ws_expenses.title = "Expenses"
        exp_headers = ["ID", "Date", "Amount ($)", "Category", "Payment Method", "Description", "Created At"]
        style_header(ws_expenses, exp_headers, fill_color="0F172A")

        # 2. Budgets Sheet
        ws_budgets = wb.create_sheet(title="Budgets")
        bud_headers = ["Category", "Monthly Budget ($)", "Updated At"]
        style_header(ws_budgets, bud_headers, fill_color="1E293B")
        today = datetime.now()
        for cat, limit in DEFAULT_BUDGETS.items():
            ws_budgets.append([cat, float(limit), today.strftime("%Y-%m-%d %H:%M")])
        auto_fit_columns(ws_budgets)

        # 3. Categories Sheet
        ws_categories = wb.create_sheet(title="Categories")
        cat_headers = ["Name", "Color", "Icon"]
        style_header(ws_categories, cat_headers, fill_color="334155")
        for c in DEFAULT_CATEGORIES:
            ws_categories.append([c["name"], c["color"], c["icon"]])
        auto_fit_columns(ws_categories)

        wb.save(target_file)
        return target_file

def get_workbook(user_id=None):
    fp = get_excel_file_path(user_id)
    init_excel(fp)
    return openpyxl.load_workbook(fp), fp

def get_categories(user_id=None):
    with _file_lock:
        wb, _ = get_workbook(user_id)
        if "Categories" not in wb.sheetnames:
            return DEFAULT_CATEGORIES
        ws = wb["Categories"]
        categories = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row and row[0]:
                categories.append({
                    "name": str(row[0]).strip(),
                    "color": str(row[1]).strip() if len(row) > 1 and row[1] else "#64748B",
                    "icon": str(row[2]).strip() if len(row) > 2 and row[2] else "tags"
                })
        return categories if categories else DEFAULT_CATEGORIES

def get_budgets(user_id=None):
    with _file_lock:
        wb, _ = get_workbook(user_id)
        if "Budgets" not in wb.sheetnames:
            return DEFAULT_BUDGETS
        ws = wb["Budgets"]
        budgets = {}
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row and row[0]:
                cat = str(row[0]).strip()
                try:
                    limit = float(row[1]) if len(row) > 1 and row[1] is not None else 0.0
                except (ValueError, TypeError):
                    limit = 0.0
                budgets[cat] = limit
        return budgets

def save_budgets(new_budgets, user_id=None):
    with _file_lock:
        wb, fp = get_workbook(user_id)
        if "Budgets" in wb.sheetnames:
            del wb["Budgets"]
        ws = wb.create_sheet(title="Budgets")
        style_header(ws, ["Category", "Monthly Budget ($)", "Updated At"], fill_color="1E293B")
        now_str = datetime.now().strftime("%Y-%m-%d %H:%M")
        for cat, limit in new_budgets.items():
            ws.append([cat, float(limit), now_str])
        auto_fit_columns(ws)
        wb.save(fp)
        return True

def get_expenses(start_date=None, end_date=None, category=None, search=None, payment_method=None, month=None, user_id=None):
    with _file_lock:
        wb, _ = get_workbook(user_id)
        ws = wb["Expenses"]
        expenses = []

        # Resolve target month filter
        target_month = str(month).strip() if month is not None else None
        if target_month and target_month.lower() in ['all', '']:
            target_month = None
        elif target_month and target_month.lower() == 'auto':
            # Auto-detect active month
            all_dates = []
            for row in ws.iter_rows(min_row=2, max_col=2, values_only=True):
                if row and row[1]:
                    d = str(row[1]).strip()
                    if isinstance(row[1], datetime):
                        d = row[1].strftime("%Y-%m-%d")
                    if len(d) >= 7:
                        all_dates.append(d[:7])
            now_mo = datetime.now().strftime("%Y-%m")
            if now_mo in all_dates:
                target_month = now_mo
            elif all_dates:
                target_month = sorted(list(set(all_dates)), reverse=True)[0]
            else:
                target_month = None

        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or row[0] is None:
                continue
            
            exp_id = int(row[0]) if isinstance(row[0], (int, float)) else str(row[0])
            date_val = str(row[1]) if row[1] is not None else ""
            if isinstance(row[1], datetime):
                date_val = row[1].strftime("%Y-%m-%d")
            
            try:
                amount_val = float(row[2]) if row[2] is not None else 0.0
            except (ValueError, TypeError):
                amount_val = 0.0
                
            cat_val = str(row[3]) if row[3] is not None else "Miscellaneous"
            pay_val = str(row[4]) if len(row) > 4 and row[4] is not None else "Other"
            desc_val = str(row[5]) if len(row) > 5 and row[5] is not None else ""
            created_val = str(row[6]) if len(row) > 6 and row[6] is not None else ""
            
            # Filtering
            if target_month and not date_val.startswith(target_month):
                continue
            if category and category.lower() != 'all' and cat_val.lower() != category.lower():
                continue
            if payment_method and payment_method.lower() != 'all' and pay_val.lower() != payment_method.lower():
                continue
            if start_date and date_val < start_date:
                continue
            if end_date and date_val > end_date:
                continue
            if search:
                search_lower = search.lower()
                if search_lower not in desc_val.lower() and search_lower not in cat_val.lower() and search_lower not in pay_val.lower():
                    continue

            expenses.append({
                "id": exp_id,
                "date": date_val,
                "amount": round(amount_val, 2),
                "category": cat_val,
                "payment_method": pay_val,
                "description": desc_val,
                "created_at": created_val
            })
            
        # Sort by date desc, id desc
        expenses.sort(key=lambda x: (x["date"], x["id"]), reverse=True)
        return expenses

def add_expense(date, amount, category, payment_method, description, user_id=None):
    with _file_lock:
        wb, fp = get_workbook(user_id)
        ws = wb["Expenses"]
        
        # Determine next ID
        max_id = 0
        for row in ws.iter_rows(min_row=2, max_col=1, values_only=True):
            if row and row[0] is not None:
                try:
                    val = int(row[0])
                    if val > max_id:
                        max_id = val
                except (ValueError, TypeError):
                    pass
        new_id = max_id + 1
        now_str = datetime.now().strftime("%Y-%m-%d %H:%M")
        
        ws.append([new_id, str(date), float(amount), str(category), str(payment_method), str(description), now_str])
        auto_fit_columns(ws)
        wb.save(fp)
        return {
            "id": new_id,
            "date": str(date),
            "amount": float(amount),
            "category": str(category),
            "payment_method": str(payment_method),
            "description": str(description),
            "created_at": now_str
        }

def bulk_add_expenses(items, user_id=None):
    with _file_lock:
        wb, fp = get_workbook(user_id)
        ws = wb["Expenses"]
        
        # Determine starting ID
        max_id = 0
        for row in ws.iter_rows(min_row=2, max_col=1, values_only=True):
            if row and row[0] is not None:
                try:
                    val = int(row[0])
                    if val > max_id:
                        max_id = val
                except (ValueError, TypeError):
                    pass
                    
        now_str = datetime.now().strftime("%Y-%m-%d %H:%M")
        added_records = []
        
        for item in items:
            max_id += 1
            date_val = str(item.get("date", datetime.now().strftime("%Y-%m-%d")))
            amt_val = float(item.get("amount", 0.0))
            cat_val = str(item.get("category", "Miscellaneous"))
            pay_val = str(item.get("payment_method", "Amex Card"))
            desc_val = str(item.get("description", "Purchase"))
            
            ws.append([max_id, date_val, amt_val, cat_val, pay_val, desc_val, now_str])
            added_records.append({
                "id": max_id,
                "date": date_val,
                "amount": amt_val,
                "category": cat_val,
                "payment_method": pay_val,
                "description": desc_val,
                "created_at": now_str
            })
            
        auto_fit_columns(ws)
        wb.save(fp)
        return added_records

def update_expense(expense_id, date, amount, category, payment_method, description, user_id=None):
    with _file_lock:
        wb, fp = get_workbook(user_id)
        ws = wb["Expenses"]
        found = False
        target_row_idx = None
        
        for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=False), start=2):
            cell_val = row[0].value
            if cell_val is not None and str(cell_val) == str(expense_id):
                target_row_idx = idx
                row[1].value = str(date)
                row[2].value = float(amount)
                row[3].value = str(category)
                if len(row) > 4: row[4].value = str(payment_method)
                if len(row) > 5: row[5].value = str(description)
                found = True
                break
                
        if found:
            auto_fit_columns(ws)
            wb.save(fp)
            return True
        return False

def delete_expense(expense_id, user_id=None):
    with _file_lock:
        wb, fp = get_workbook(user_id)
        ws = wb["Expenses"]
        target_row_idx = None
        for idx, row in enumerate(ws.iter_rows(min_row=2, max_col=1, values_only=True), start=2):
            if row and row[0] is not None and str(row[0]) == str(expense_id):
                target_row_idx = idx
                break
        if target_row_idx is not None:
            ws.delete_rows(target_row_idx, 1)
            wb.save(fp)
            return True
        return False

def clear_all_expenses(month=None, user_id=None):
    with _file_lock:
        wb, fp = get_workbook(user_id)
        if "Expenses" not in wb.sheetnames:
            return 0
        ws = wb["Expenses"]
        
        # If month is specific (e.g. "2026-08")
        if month and month.lower() not in ['all', 'auto', '']:
            deleted_count = 0
            # Iterate backwards to safely delete matching rows
            for row_idx in range(ws.max_row, 1, -1):
                date_cell = ws.cell(row=row_idx, column=2).value
                if date_cell and str(date_cell).strip().startswith(month):
                    ws.delete_rows(row_idx, 1)
                    deleted_count += 1
            auto_fit_columns(ws)
            wb.save(fp)
            return deleted_count
        else:
            # Delete all rows across all months
            del wb["Expenses"]
            ws = wb.create_sheet(title="Expenses", index=0)
            exp_headers = ["ID", "Date", "Amount ($)", "Category", "Payment Method", "Description", "Created At"]
            style_header(ws, exp_headers, fill_color="0F172A")
            auto_fit_columns(ws)
            wb.save(fp)
            return -1

def get_summary_stats(month=None, user_id=None):
    expenses = get_expenses(user_id=user_id)
    budgets = get_budgets(user_id=user_id)
    categories_list = get_categories(user_id=user_id)
    cat_meta = {c["name"]: c for c in categories_list}
    
    now = datetime.now()
    current_calendar_month = now.strftime("%Y-%m")
    
    # 1. Discover all months present in expenses
    unique_months = sorted(list(set(e["date"][:7] for e in expenses if len(e.get("date", "")) >= 7)), reverse=True)
    if current_calendar_month not in unique_months:
        unique_months.insert(0, current_calendar_month)

    # 2. Determine target active month
    if month and month.lower() != 'auto' and month.lower() != 'all':
        active_month = month
    else:
        # Auto-detection:
        # Check if current calendar month has expenses
        curr_has_expenses = any(e["date"].startswith(current_calendar_month) for e in expenses)
        if curr_has_expenses:
            active_month = current_calendar_month
        elif expenses:
            active_month = expenses[0]["date"][:7]
        else:
            active_month = current_calendar_month

    # 3. Calculate previous month prefix relative to active_month
    try:
        y, m = map(int, active_month.split('-'))
        dt_active = datetime(y, m, 1)
        dt_prev = (dt_active - timedelta(days=1)).replace(day=1)
        prev_month_prefix = dt_prev.strftime("%Y-%m")
    except Exception:
        prev_month_prefix = ""

    # 4. Filter expenses
    total_all_time = sum(e["amount"] for e in expenses)
    if month and month.lower() == 'all':
        month_expenses = expenses
        prev_month_expenses = []
        month_display_title = "All Time"
    else:
        month_expenses = [e for e in expenses if e["date"].startswith(active_month)]
        prev_month_expenses = [e for e in expenses if e["date"].startswith(prev_month_prefix)]
        try:
            month_display_title = datetime.strptime(active_month, "%Y-%m").strftime("%B %Y")
        except Exception:
            month_display_title = active_month
    
    month_spend = sum(e["amount"] for e in month_expenses)
    prev_month_spend = sum(e["amount"] for e in prev_month_expenses)
    
    # Total monthly budget
    total_monthly_budget = sum(budgets.values())
    remaining_budget = max(0.0, total_monthly_budget - month_spend)
    budget_usage_pct = round((month_spend / total_monthly_budget * 100), 1) if total_monthly_budget > 0 else 0
    
    # Daily average
    if active_month == current_calendar_month:
        days_count = max(1, now.day)
    else:
        try:
            y, m = map(int, active_month.split('-'))
            import calendar
            days_count = calendar.monthrange(y, m)[1]
        except Exception:
            days_count = 30
    daily_avg_spend = round(month_spend / days_count, 2)
    
    # Category Breakdown
    category_totals = {}
    for e in month_expenses:
        cat = e["category"]
        category_totals[cat] = category_totals.get(cat, 0.0) + e["amount"]
        
    category_breakdown = []
    for cat_name, amt in category_totals.items():
        meta = cat_meta.get(cat_name, {"color": "#64748B", "icon": "tags"})
        budget_limit = budgets.get(cat_name, 0.0)
        pct_of_total = round((amt / month_spend * 100), 1) if month_spend > 0 else 0
        pct_of_budget = round((amt / budget_limit * 100), 1) if budget_limit > 0 else 0
        category_breakdown.append({
            "category": cat_name,
            "amount": round(amt, 2),
            "percentage": pct_of_total,
            "budget": budget_limit,
            "budget_usage_pct": pct_of_budget,
            "color": meta["color"],
            "icon": meta["icon"]
        })
    category_breakdown.sort(key=lambda x: x["amount"], reverse=True)
    
    # Budget vs Actual comparison
    budget_comparison = []
    for cat_name, limit in budgets.items():
        actual = category_totals.get(cat_name, 0.0)
        meta = cat_meta.get(cat_name, {"color": "#64748B", "icon": "tags"})
        usage = round((actual / limit * 100), 1) if limit > 0 else 0
        budget_comparison.append({
            "category": cat_name,
            "budget": round(limit, 2),
            "actual": round(actual, 2),
            "remaining": round(max(0, limit - actual), 2),
            "usage_pct": usage,
            "is_over": actual > limit,
            "color": meta["color"],
            "icon": meta["icon"]
        })
    budget_comparison.sort(key=lambda x: x["actual"], reverse=True)

    # Timeline trend for the active month (or last 30 days if current)
    timeline_days = {}
    if active_month == current_calendar_month:
        for i in range(29, -1, -1):
            day_str = (now - timedelta(days=i)).strftime("%Y-%m-%d")
            timeline_days[day_str] = 0.0
    else:
        try:
            y, m = map(int, active_month.split('-'))
            import calendar
            num_days = calendar.monthrange(y, m)[1]
            for day in range(1, num_days + 1):
                day_str = f"{y:04d}-{m:02d}-{day:02d}"
                timeline_days[day_str] = 0.0
        except Exception:
            for i in range(29, -1, -1):
                day_str = (now - timedelta(days=i)).strftime("%Y-%m-%d")
                timeline_days[day_str] = 0.0

    for e in month_expenses:
        if e["date"] in timeline_days:
            timeline_days[e["date"]] += e["amount"]
            
    timeline_labels = [d[5:] for d in timeline_days.keys()]  # MM-DD
    timeline_values = [round(val, 2) for val in timeline_days.values()]
    
    # Payment Method breakdown
    payment_method_totals = {}
    for e in month_expenses:
        p = e.get("payment_method") or "Other"
        payment_method_totals[p] = payment_method_totals.get(p, 0.0) + e["amount"]
        
    top_category = category_breakdown[0]["category"] if category_breakdown else "None"
    top_category_amount = category_breakdown[0]["amount"] if category_breakdown else 0.0
    
    # Month-over-month growth percentage
    if prev_month_spend > 0:
        mom_growth_pct = round(((month_spend - prev_month_spend) / prev_month_spend) * 100, 1)
    else:
        mom_growth_pct = 0.0

    # Format available months list for frontend dropdown
    available_months = []
    for ym in unique_months:
        try:
            lbl = datetime.strptime(ym, "%Y-%m").strftime("%B %Y")
        except Exception:
            lbl = ym
        available_months.append({
            "value": ym,
            "label": lbl,
            "is_active": (ym == active_month)
        })

    return {
        "active_month": active_month,
        "active_month_label": month_display_title,
        "available_months": available_months,
        "current_month_spend": round(month_spend, 2),
        "prev_month_spend": round(prev_month_spend, 2),
        "mom_growth_pct": mom_growth_pct,
        "total_monthly_budget": round(total_monthly_budget, 2),
        "remaining_budget": round(remaining_budget, 2),
        "budget_usage_pct": budget_usage_pct,
        "daily_avg_spend": daily_avg_spend,
        "total_all_time": round(total_all_time, 2),
        "total_transactions_count": len(expenses),
        "current_month_count": len(month_expenses),
        "top_category": top_category,
        "top_category_amount": round(top_category_amount, 2),
        "category_breakdown": category_breakdown,
        "budget_comparison": budget_comparison,
        "timeline": {
            "labels": timeline_labels,
            "values": timeline_values,
            "raw_dates": list(timeline_days.keys())
        },
        "payment_methods": payment_method_totals
    }
